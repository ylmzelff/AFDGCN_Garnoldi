"""
export_ildem_30days.py
======================
Kayseri Büyükşehir Belediyesi API'sinden İldem bölgesi için
bugünden başlayarak geriye doğru 30 günlük veri çekip
ocak_gerçek.xlsx ile aynı formatta Excel dosyası oluşturur.

Kullanım:
    python scripts/export_ildem_30days.py
    python scripts/export_ildem_30days.py --days 30 --output ildem_son30gun.xlsx
"""

import sys
import time
import logging
import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import requests
import urllib3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# SSL uyarılarını kapat (self-signed certificate)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# Kayseri Belediye API yapılandırması
# ─────────────────────────────────────────────────────────────────────────────

BELEDIYE_BASE_URL = "https://10.50.234.18"

BELEDIYE_COOKIES = {
    ".AspNetCore.Antiforgery.708PDUFCWo4": (
        "CfDJ8B9JLEL67llGtXY4pwUvuwygNgpVdSBpqo58gIsraZTkJoO9l_vmK9fPZTbFz9qKHy"
        "hKGyWbMEZRP1jjNbO5KB2fA7vT9ZHH8dNFc65mUPzQ5zFiCCa7UuZGI1CoNMKwR3C2rkwO"
        "TkJ6IXeGRRqoXG4"
    ),
    ".AspNetCore.Identity.Application": (
        "CfDJ8B9JLEL67llGtXY4pwUvuwzENkFdMmSSt0X8NUoZ7onLJ3F_hVrpfwtdQjnfgKki_oV"
        "aTGU_agTPcGk1iB3mf2PHjYJMHRyZISK5weVMU3dTmgJyfgIeSVFLqnSdWo0jNa-O3gxmzv"
        "FiMZ-hNI-v_ufzg4mL48kU-mWAFq9RzF2xdEyehBrTMA-QKo7fW7pcIc2sfCsPkzDuF6gcV"
        "9u2bI4Jt_CgO9twe18_VAl8Jh8R0PusijTPJ73Qgiepg0XZwuYitMs9IL7OOkkpPU6Odx_p"
        "0eBJ6uIZ4aluDv4zjGb98jQUi3sbyYW6gFWU1WaYq01ehbQ7OD8lLUhEdYVD8aC-nOtMtha"
        "qm4jkoOi5xqAK-3aeICcSYkg92U9R1-1cbcSk4YDz6aqLKWVoLsCzi7zfJpYqrHUu-FIx1l"
        "z_7pbMFWZWcW4EwBQw6X-_y8wj7pWLFmTthMTfuZwI7WKHglH-N0WVa3kKWyEU0TD0mhcRc"
        "idHViII51Wq2SpVug98M2_mz63UA_NbcR_AeNNBFwIwBSSGhXmSzoQWmriPb4HcdHNZqptb1"
        "IwLsfaqddT_560sdoSL02CWI7Z-fLGpfSFCVsyOUdSqenUXFNEDFuKOckbfPqKnSg4VlfcDO"
        "7_hrQyt0ZqKGNxijpwb9TGctcUuWgY7-nM9dUx2zgHl_LTx3ns2g2UE1QMVG9DTVlyXvQz7"
        "_jaDhAbrdqvxVOJ9PKzVL2wCbbBhwRaMPtzm31lAMpz2tvi6glH183LQMYf5Z3u0o7uoaO_w"
        "sTBQIuHM-X0pkzNaJ7LL_s2cb6d96XOlJw6GxwudHsZZl0xs6g8p-wd96JdQ5-qlFyPBN4nU"
        "gsuJKL-YSoCNlwF44VUnqHj-YSoCNlwF44VUnqHj"
    ),
}

# ─────────────────────────────────────────────────────────────────────────────
# Kavşak → sayfa adı + kol adları
# ─────────────────────────────────────────────────────────────────────────────

# Kol sırası ve isimleri (yoksa o sütun eklenmez)
JUNCTION_CONFIG: Dict[int, Dict] = {
    89: {
        "sheet": "Gesi",
        "arms": {
            "A": "SİVAS BULVARI-SİVAS YÖNÜ",
            "B": "GESİ CAD.",
            "C": "SİVAS BULV - ŞEHİR MERKEZİ",
            "D": "381. SOKAK",
        },
    },
    187: {
        "sheet": "Serkent",
        "arms": {
            "A": "822. SK",
            "B": "GESİ CAD. DOĞU",
            "C": "KOCASİNAN CAD.",
            "D": "GESİ CAD. BATI",
        },
    },
    95: {
        "sheet": "Beyazşehir",
        "arms": {
            "A": "OSMAN ÖZCAN CAD.",
            "B": "GESİ CAD DOĞU",
            "C": "MARKETLER ÇIKIŞ",
            "D": "GESİ CAD BATI",
        },
    },
    121: {
        "sheet": "Toki",
        "arms": {
            "A": "OSMAN ÖZCAN CAD.",
            "B": "GESİ CAD DOĞU",
            "C": "MARKETLER ÇIKIŞ",
            "D": "GESİ CAD BATI",
        },
    },
    184: {
        "sheet": "İldem 1",
        "arms": {
            "A": "DİNÇER SOKAK",
            "B": "ALPARSLANTÜRKEŞ BUL.",
            "D": "GESİ CAD",
        },
    },
    188: {
        "sheet": "İldem 2",
        "arms": {
            "A": "DİNÇER SOKAK KUZEY",
            "B": "HANEDAN SOKAK.",
            "C": "DİNÇER SOKAK GÜNEY",
            "D": "FETİH SOKAK",
        },
    },
    117: {
        "sheet": "İldem 3",
        "arms": {
            "A": "DÜNDAR TAŞER CAD. KUZEY",
            "C": "DÜNDAR TAŞER CAD. GÜNEY",
            "D": "HANEDAN SOKAK.",
        },
    },
    192: {
        "sheet": "İldem 4",
        "arms": {
            "A": "YAVUZSULTAN SELİM CAD. KUZEY",
            "B": "VATAN SOKAK",
            "C": "YAVUZSULTAN SELİM CAD. GÜNEY",
            "D": "ORKUN SOKAK",
        },
    },
    194: {
        "sheet": "İldem 5",
        "arms": {
            "A": "S.A BEDUK CAD. KUZEY",
            "B": "VATAN SOKAK BATI",
            "C": "S.A BEDUK CAD. GÜNEY",
            "D": "VATAN SOKAK DOĞU",
        },
    },
}

# Kavşak işlem sırası (orijinal xlsx ile aynı)
JUNCTION_ORDER = [89, 187, 95, 121, 184, 188, 117, 192, 194]


# ─────────────────────────────────────────────────────────────────────────────
# Belediye veri çekici
# ─────────────────────────────────────────────────────────────────────────────

class BelediyeFetcher:
    """Kayseri Belediye API'sinden 10 dakikalık araç sayısı verisi çeker."""

    def __init__(self):
        self.session = requests.Session()
        self.session.verify = False
        self.session.cookies.update(BELEDIYE_COOKIES)
        self.session.headers.update(
            {
                "accept": "*/*",
                "cache-control": "no-cache",
                "pragma": "no-cache",
                "referer": f"{BELEDIYE_BASE_URL}/Rapor",
                "x-requested-with": "XMLHttpRequest",
                "user-agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0 Safari/537.36"
                ),
            }
        )

    def fetch(
        self, junction_id: int, date_str: str
    ) -> Optional[List[Dict]]:
        """
        Parameters
        ----------
        junction_id : int
        date_str    : str  -- "YYYY-MM-DD"

        Returns
        -------
        list of dicts where each dict is one arm/edge with keys:
            edge_direction ('A'/'B'/'C'/'D'), edge_name, '0'..'143'
        None on failure.
        """
        url = f"{BELEDIYE_BASE_URL}/Rapor/GetDakikalikRapor"
        params = {"id": junction_id, "date": f"{date_str}T00:00", "wa": 0}
        try:
            r = self.session.get(url, params=params, timeout=30)
            if r.status_code != 200:
                log.warning(
                    "HTTP %s  kavşak=%s  tarih=%s", r.status_code, junction_id, date_str
                )
                return None
            payload = r.json()
            if isinstance(payload, dict):
                return payload.get("data") or payload.get("results") or payload
            return payload if isinstance(payload, list) else None
        except Exception as exc:
            log.warning("Hata kavşak=%s tarih=%s: %s", junction_id, date_str, exc)
            return None


# ─────────────────────────────────────────────────────────────────────────────
# Yardımcı: API verisinden kol bazlı 10-dakikalık diziye dönüştür
# ─────────────────────────────────────────────────────────────────────────────

def extract_arm_series(
    raw_data: List[Dict], arms: Dict[str, str]
) -> Dict[str, List[float]]:
    """
    raw_data → {arm_direction: [v0, v1, ..., v143]}
    Sadece config'de tanımlı kolları alır.
    """
    direction_map: Dict[str, Dict] = {}
    for edge in raw_data:
        d = str(edge.get("edge_direction", "")).strip().upper()
        if d in arms:
            direction_map[d] = edge

    result: Dict[str, List[float]] = {}
    for arm_dir in arms:
        edge = direction_map.get(arm_dir)
        if edge:
            result[arm_dir] = [float(edge.get(str(i), 0) or 0) for i in range(144)]
        else:
            result[arm_dir] = [0.0] * 144
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Excel sayfası oluştur
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)


def write_sheet(ws, junction_id: int, all_days: Dict[str, Dict[str, List[float]]]):
    """
    all_days : {date_str: {arm_dir: [v0..v143]}}
    """
    cfg = JUNCTION_CONFIG[junction_id]
    arms = cfg["arms"]
    arm_dirs = list(arms.keys())

    # ── Başlık satırı (satır 1) ──────────────────────────────────────────────
    ws.cell(1, 1, "KBB TRAFİK YÖNETİM VE RAPORLAMA").font = TITLE_FONT
    # Satır 2 boş

    # ── Sütun başlıkları (satır 3) ───────────────────────────────────────────
    headers = ["Tarih", "Saat"] + [f"{d} - {arms[d]}" for d in arm_dirs] + ["Toplam"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(3, col_idx, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ── Veri satırları ───────────────────────────────────────────────────────
    row = 4
    for date_str in sorted(all_days.keys()):
        arm_series = all_days[date_str]
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")

        for idx in range(144):
            hr = idx // 6
            mn = (idx % 6) * 10
            time_str = f"{hr:02d}:{mn:02d}"

            vals = [arm_series[d][idx] for d in arm_dirs]
            total = sum(vals)

            ws.cell(row, 1, date_obj).number_format = "DD.MM.YYYY"
            ws.cell(row, 2, time_str)
            for col_offset, v in enumerate(vals, start=3):
                ws.cell(row, col_offset, int(round(v)))
            ws.cell(row, 3 + len(vals), int(round(total)))
            row += 1

    # ── Sütun genişlikleri ───────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 8
    for col_idx in range(3, 3 + len(arm_dirs) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 28


# ─────────────────────────────────────────────────────────────────────────────
# Ana işlem
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="İldem bölgesi 30 günlük Excel export")
    parser.add_argument("--days", type=int, default=30, help="Kaç gün geriye gidilsin (varsayılan: 30)")
    parser.add_argument(
        "--output",
        default="ildem_son30gun.xlsx",
        help="Çıktı dosyası (varsayılan: ildem_son30gun.xlsx)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="İstekler arası bekleme süresi saniye cinsinden (varsayılan: 0.5)",
    )
    args = parser.parse_args()

    today = datetime.now().date()
    date_list = [
        (today - timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(args.days - 1, -1, -1)  # eskiden yeniye sıralı
    ]

    log.info("Tarih aralığı: %s → %s  (%d gün)", date_list[0], date_list[-1], len(date_list))
    log.info("Kavşak sayısı: %d", len(JUNCTION_ORDER))
    log.info("Toplam API çağrısı: ~%d", len(date_list) * len(JUNCTION_ORDER))

    fetcher = BelediyeFetcher()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # varsayılan boş sayfayı sil

    for junction_id in JUNCTION_ORDER:
        cfg = JUNCTION_CONFIG[junction_id]
        sheet_name = cfg["sheet"]
        arms = cfg["arms"]

        log.info("─── %s (kavşak %d) ───", sheet_name, junction_id)
        ws = wb.create_sheet(title=sheet_name)

        all_days: Dict[str, Dict[str, List[float]]] = {}

        for date_str in date_list:
            log.info("  ↳ %s ...", date_str)
            raw = fetcher.fetch(junction_id, date_str)

            if raw:
                all_days[date_str] = extract_arm_series(raw, arms)
                log.info("    ✓ %d kol, 144 zaman dilimi", len(arms))
            else:
                # Boş gün: sıfır dizi
                all_days[date_str] = {d: [0.0] * 144 for d in arms}
                log.warning("    ✗ Veri alınamadı – sıfır ile dolduruldu")

            time.sleep(args.delay)

        write_sheet(ws, junction_id, all_days)
        log.info("  ✔ Sayfa yazıldı: %s", sheet_name)

    # Kaydet
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = Path(__file__).parent.parent / output_path

    wb.save(str(output_path))
    log.info("═" * 60)
    log.info("✅  Dosya kaydedildi: %s", output_path)
    log.info("    Boyut: %.1f KB", output_path.stat().st_size / 1024)


if __name__ == "__main__":
    main()
