"""
Excel dosyasını düzeltme scripti:
1. Mayıs (5. ay) verilerini çıkar
2. 18 Nisan 2026 verilerini ekle
"""
import json
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

# --- Kaynak dosyalar ---
BASE = Path(r"C:\Users\lenovo\Desktop\Yeni klasör\Yeni_klasör\projects\AFDGCN_Garnoldi")
XLSX_PATH = BASE / "İLDEM_1aylik_ayrı_kavşaklar.xlsx"
TRANSCRIPT_PATH = Path(r"C:\Users\lenovo\AppData\Roaming\Code\User\workspaceStorage\84c4f0a74186b4ee9d62d444afa88ba0\GitHub.copilot-chat\transcripts\e78c63e7-96fd-4c17-bb0b-eeb80255148e.jsonl")

# --- JSON verisini transcript'tan çek ---
print("Transcript'tan April 18 verisi yükleniyor...")
with open(TRANSCRIPT_PATH, "r", encoding="utf-8") as f:
    lines = f.readlines()

apr18_data = None
for line in lines:
    msg = json.loads(line)
    if msg.get("type") == "user.message":
        content = msg["data"]["content"]
        if '"tarih": "2026-04-18"' in content:
            start = content.index("{")
            # Brace-balanced JSON extraction
            depth = 0
            end = start
            for i, ch in enumerate(content[start:], start):
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        end = i + 1
                        break
            apr18_data = json.loads(content[start:end])
            break

if apr18_data is None:
    raise RuntimeError("April 18 verisi transcript'ta bulunamadı!")

print(f"April 18 kayıt sayısı: {len(apr18_data['data'])}")

# --- April 18 verilerini iç yapıya dönüştür ---
# {intersectionId: {edgeDir: {saat: count}}}
apr18_by_id = {}
for entry in apr18_data["data"]:
    iid = entry["intersectionId"]
    ed = entry["edgeDirection"]
    if iid not in apr18_by_id:
        apr18_by_id[iid] = {}
    apr18_by_id[iid][ed] = entry["saatlikVeriler"]

print("Kesişim IDs:", sorted(apr18_by_id.keys()))

# --- Sheet -> (intersectionId, [kolumlar sırası]) mapping ---
# Kolumlar, başlık satırındaki sıraya göre A,B,C,D yön harfleri
SHEET_MAPPING = {
    "Gesi":       (89,  ["A", "B", "C", "D"]),
    "Serkent":    (187, ["A", "B", "C", "D"]),
    "Beyazşehir": (95,  ["A", "B", "C", "D"]),
    "Toki":       (121, ["A", "B", "C", "D"]),  # JSON'da yok, 0 kullanılacak
    "İldem 1":    (184, ["A", "B", "D"]),         # C kolonu yok
    "İldem 2":    (188, ["A", "B", "C", "D"]),
    "İldem 3":    (117, ["A", "C", "D"]),          # B kolonu yok
    "İldem 4":    (192, ["A", "B", "C", "D"]),
    "İldem 5":    (194, ["A", "B", "C", "D"]),
}

# 144 zaman dilimi üret (00:00 - 23:50, 10 dk)
TIME_SLOTS = []
for h in range(24):
    for m in range(0, 60, 10):
        TIME_SLOTS.append(f"{h:02d}:{m:02d}")

APR18_DATE = "18.04.2026"
MAY_SUFFIX = ".05."  # Bu string tarih içindeyse Mayıs demektir

# --- Excel'i yükle (tam mod) ---
print(f"Excel dosyası yükleniyor: {XLSX_PATH}")
wb = openpyxl.load_workbook(XLSX_PATH)

for sheet_name, (iid, dirs) in SHEET_MAPPING.items():
    print(f"\nSheet işleniyor: {sheet_name} (id={iid}, dirs={dirs})")
    ws = wb[sheet_name]

    # Tüm satırları oku
    all_rows = list(ws.iter_rows(values_only=True))
    # Header: row 3 (index 2), data: row 4+ (index 3+)
    header_rows = all_rows[:3]   # ilk 3 satır (2 boş + başlık)
    data_rows = all_rows[3:]

    # Mayıs satırlarını filtrele
    april_rows = [r for r in data_rows if r[0] is not None and MAY_SUFFIX not in str(r[0])]
    may_count = len(data_rows) - len(april_rows)
    print(f"  Mayıs satırı çıkarıldı: {may_count}")

    # April 18 satırlarını oluştur
    n_cols = len(header_rows[2])  # başlık sütun sayısı
    # Toplam kolonu her zaman son kolondur

    apr18_rows = []
    for t in TIME_SLOTS:
        row = [None] * n_cols
        row[0] = APR18_DATE
        row[1] = t

        # Her yön kolonu için değeri al
        for col_idx, direction in enumerate(dirs, start=2):
            if iid in apr18_by_id and direction in apr18_by_id[iid]:
                val = apr18_by_id[iid][direction].get(t, 0)
            else:
                val = 0
            row[col_idx] = val

        # Toplam (son kolon)
        total = sum(row[2:n_cols - 1])
        row[n_cols - 1] = total
        apr18_rows.append(tuple(row))

    print(f"  April 18 satırı eklendi: {len(apr18_rows)}")

    # Nereye ekleyeceğimizi bul: 17 Nisan'dan sonra, 19 Nisan'dan önce
    insert_idx = None
    for i, r in enumerate(april_rows):
        if r[0] is not None and r[0] == "19.04.2026":
            insert_idx = i
            break

    if insert_idx is None:
        # 19 Nisan yoksa, sona ekle
        combined = april_rows + apr18_rows
        print(f"  April 18 sona eklendi (19 Nisan bulunamadı)")
    else:
        combined = april_rows[:insert_idx] + apr18_rows + april_rows[insert_idx:]
        print(f"  April 18 {insert_idx}. konuma eklendi (19 Nisan öncesi)")

    # Sheet'i temizle (header satırları hariç)
    # Önce mevcut data satırlarını sil (satır 4'ten sona kadar)
    max_row = ws.max_row
    for row_num in range(max_row, 3, -1):
        ws.delete_rows(row_num)

    # Yeni verileri yaz
    for row_data in combined:
        ws.append(list(row_data))

    print(f"  Toplam satır yazıldı: {len(combined)}")

# --- Kaydet ---
output_path = XLSX_PATH  # Aynı dosyayı güncelle
wb.save(output_path)
print(f"\nDosya kaydedildi: {output_path}")
print("Tamamlandı!")
